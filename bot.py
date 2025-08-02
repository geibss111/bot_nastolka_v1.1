# C:\Users\Дианка\PycharmProjects\pythonProject4\bot_tennis4_beta\bot.py
import asyncio
import logging
from telegram import Bot
from telegram.error import TelegramError
from auth import login_and_save_cookies, UP_GAMES_URL
# Предполагается, что parser.py находится в той же директории
from parser import parse_html_content
import requests
import urllib3
import json
import os
from datetime import datetime, timedelta

# --- Добавленные импорты для работы с Excel ---
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# --- Конфигурация ---
# Updated token and channel ID for local run
TELEGRAM_BOT_TOKEN = "токен телеграм бота"
TELEGRAM_CHANNEL_ID = "@телеграм паблик"
# Файлы для хранения состояния (в той же папке, что и скрипт)
SENT_GAMES_FILE = "sent_games.json"
RESET_TIME_FILE = "reset_time.json"  # Сброс раз в 6 часов теперь
# --- Файл Excel для сохранения отправленных игр ---
SENT_GAMES_EXCEL_FILE = "sent_games.xlsx" # <-- Новый файл
# -------------------------------------------------
# Интервалы
CHECK_INTERVAL_SECONDS = 60  # Проверять каждую минуту
RESET_HOURS = 6  # <-- Изменено на 6 часов
# --- Настройка логирования ---
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
    handlers=[
        logging.FileHandler("bot.log"),  # Лог будет сохраняться в bot.log в той же папке
        logging.StreamHandler()  # Вывод в консоль
    ]
)
logger = logging.getLogger(__name__)
# --- Глобальные переменные для состояния ---
sent_games = set()
last_reset_time = None

# --- Функция инициализации Excel файла ---
def initialize_excel_file():
    """Создает новый Excel файл с заголовками, если он не существует."""
    if not os.path.exists(SENT_GAMES_EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Отправленные Игры"
        # Создаем заголовки
        headers = ["ПАРА", "дата и время", "Личные встречи"]
        ws.append(headers)
        # Сделаем заголовки жирными для красоты (опционально)
        for col_num, header in enumerate(headers, 1):
             cell = ws.cell(row=1, column=col_num)
             cell.font = Font(bold=True)
        try:
            wb.save(SENT_GAMES_EXCEL_FILE)
            logger.info(f"Создан новый Excel файл: {SENT_GAMES_EXCEL_FILE}")
        except Exception as e:
            logger.error(f"Ошибка создания Excel файла {SENT_GAMES_EXCEL_FILE}: {e}")
    else:
        logger.info(f"Excel файл {SENT_GAMES_EXCEL_FILE} уже существует.")

# --- Функция добавления записи в Excel ---
def append_game_to_excel(game_data):
    """Добавляет информацию об отправленной игре в Excel файл."""
    try:
        # Открываем существующий файл
        wb = load_workbook(SENT_GAMES_EXCEL_FILE)
        ws = wb.active

        # Подготавливаем данные для записи
        пара = f"{game_data['player1']} - {game_data['player2']}"
        дата_время = game_data['time'] # Используем как есть, форматирование можно добавить позже при необходимости
        личные_встречи = game_data['last_encounters_score']

        # Добавляем строку с данными
        ws.append([пара, дата_время, личные_встречи])

        # Сохраняем файл
        wb.save(SENT_GAMES_EXCEL_FILE)
        logger.info(f"Игра '{пара}' добавлена в Excel файл.")
    except Exception as e:
        logger.error(f"Ошибка добавления игры в Excel файл {SENT_GAMES_EXCEL_FILE}: {e}")

def load_sent_games():
    """Загружает список уже отправленных игр и время последнего сброса из файлов."""
    global sent_games, last_reset_time
    sent_games = set()
    if os.path.exists(SENT_GAMES_FILE):
        try:
            with open(SENT_GAMES_FILE, 'r', encoding='utf-8') as f:
                games_list = json.load(f)
                # Преобразуем списки обратно в кортежи для использования в set
                sent_games = set(tuple(game) for game in games_list)
            logger.info(f"Загружено {len(sent_games)} отправленных игр из файла.")
        except Exception as e:
            logger.error(f"Ошибка загрузки отправленных игр: {e}")
            sent_games = set()  # Сбрасываем на пустое множество в случае ошибки
    # Загрузка времени последнего сброса
    if os.path.exists(RESET_TIME_FILE):
        try:
            with open(RESET_TIME_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                last_reset_time = datetime.fromisoformat(data['last_reset'])
            logger.info(f"Время последнего сброса: {last_reset_time}")
        except Exception as e:
            logger.error(f"Ошибка загрузки времени сброса: {e}")
            # Если файл поврежден или ошибка, считаем, что сброс нужен
            last_reset_time = datetime.min
    else:
        # Если файла нет, считаем, что первый запуск
        last_reset_time = datetime.min
    # Если last_reset_time все еще None (например, после ошибки и datetime.min),
    # установим его в datetime.min, чтобы гарантировать сброс при первом запуске или ошибке
    if last_reset_time is None:
        last_reset_time = datetime.min

def save_sent_games():
    """Сохраняет текущий список отправленных игр в файл."""
    try:
        # Преобразуем кортежи в списки для сериализации в JSON
        games_list = [list(game) for game in sent_games]
        with open(SENT_GAMES_FILE, 'w', encoding='utf-8') as f:
            json.dump(games_list, f, ensure_ascii=False, indent=2)
        logger.debug(f"Сохранено {len(sent_games)} отправленных игр в файл.")
    except Exception as e:
        logger.error(f"Ошибка сохранения отправленных игр: {e}")

def save_reset_time():
    """Сохраняет время последнего сброса в файл."""
    try:
        with open(RESET_TIME_FILE, 'w', encoding='utf-8') as f:
            json.dump({'last_reset': last_reset_time.isoformat()}, f, ensure_ascii=False, indent=2)
        logger.debug(f"Время сброса сохранено: {last_reset_time}")
    except Exception as e:
        logger.error(f"Ошибка сохранения времени сброса: {e}")

def should_reset_sent_games():
    """Проверяет, нужно ли сбросить список отправленных игр."""
    global last_reset_time
    now = datetime.now()
    # Проверяем, прошло ли RESET_HOURS с момента последнего сброса
    if now - last_reset_time >= timedelta(hours=RESET_HOURS):
        logger.info(f"{RESET_HOURS} часов прошло. Сбрасываем список отправленных игр.")
        last_reset_time = now
        save_reset_time()  # Сохраняем новое время сброса
        return True
    return False

def is_game_already_sent(game):
    """Проверяет, было ли сообщение об этой игре уже отправлено."""
    # Ключ включает игроков, счет, лигу и время (если доступна)
    game_key = (
        game['player1'],
        game['player2'],
        game['last_encounters_score'],
        game.get('league', 'Неизвестная лига'),  # Используем значение по умолчанию
        game.get('time', 'N/A'),  # Добавляем время для различия
        game.get('wins_player1', 0),  # Добавляем счет для различия
        game.get('wins_player2', 0)
    )
    return game_key in sent_games

def mark_game_as_sent(game):
    """Добавляет игру в список отправленных."""
    game_key = (
        game['player1'],
        game['player2'],
        game['last_encounters_score'],
        game.get('league', 'Неизвестная лига'),
        game.get('time', 'N/A'),
        game.get('wins_player1', 0),
        game.get('wins_player2', 0)
    )
    sent_games.add(game_key)
    save_sent_games()  # Сохраняем сразу после добавления

async def send_telegram_message(message):
    """Асинхронно отправляет сообщение в Telegram канал БЕЗ предпоказа ссылок."""
    bot = Bot(token=TELEGRAM_BOT_TOKEN)
    try:
        # Отправляем сообщение с отключенным link preview
        await bot.send_message(
            chat_id=TELEGRAM_CHANNEL_ID,
            text=message,
            parse_mode="HTML",
            disable_web_page_preview=True  # Отключаем предпоказ ссылок
        )
        logger.info("Сообщение успешно отправлено в Telegram (без предпоказа).")
    except TelegramError as e:
        logger.error(f"Ошибка отправки сообщения в Telegram: {e}")

async def check_and_send_games():
    """Основная логика проверки игр и отправки уведомлений."""
    global sent_games
    logger.info("Начинаем проверку предстоящих игр...")
    # Проверка необходимости сброса
    if should_reset_sent_games():
        sent_games.clear()
        save_sent_games()  # Сохраняем пустой список
        logger.info("Список отправленных игр очищен.")
    # Авторизация и получение сессии
    session = login_and_save_cookies()
    if not session:
        logger.error("Не удалось получить авторизованную сессию. Проверка пропущена.")
        return
    try:
        logger.info(f"Выполняем запрос к: {UP_GAMES_URL}")
        # Отключаем SSL-верификацию и устанавливаем таймаут
        response = session.get(UP_GAMES_URL, verify=False, timeout=30)
        logger.info(f"Статус ответа: {response.status_code}")
        if response.status_code == 200:
            html_content = response.text
            logger.info(f"Получено {len(html_content)} символов HTML.")
            # Парсим HTML для извлечения игр
            games = parse_html_content(html_content)
            logger.info(f"Найдено и распаршено {len(games)} игр.")
            new_games_count = 0
            for i, game in enumerate(games):
                logger.debug(f"Проверяем игру {i + 1}: {game}")
                # Проверяем критерий суммы очков
                if game["total_last_encounters_score"] >= 10:
                    # Проверяем, не отправляли ли мы уже это сообщение
                    if not is_game_already_sent(game):
                        # --- Формируем сообщение с HTML-разметкой и РЕКОМЕНДАЦИЕЙ ---
                        league_info = game.get('league', 'Неизвестная лига')
                        # Убедимся, что время отформатировано правильно (пробел между датой и временем)
                        formatted_time = game['time']
                        if formatted_time != "N/A" and ' ' not in formatted_time and len(
                                formatted_time) == 11:  # e.g., "26.0721:20"
                            formatted_time = f"{formatted_time[:5]} {formatted_time[5:]}"  # "26.07 21:20"
                        # --- Логика рекомендации ставки ---
                        recommendation = "Ставка: Не определена"
                        try:
                            wins_p1 = game.get('wins_player1', 0)
                            wins_p2 = game.get('wins_player2', 0)
                            # Проверяем, можно ли дать рекомендацию (счет не 0:0)
                            if wins_p1 == 0 and wins_p2 == 0:
                                recommendation = "Ставка: Рекомендация недоступна (0:0)"
                            elif wins_p1 < wins_p2:
                                # У Игрока 1 меньше побед -> Ставка П1 в первом сете
                                recommendation = f"Ставка: П1 в первом сете"
                            elif wins_p2 < wins_p1:
                                # У Игрока 2 меньше побед -> Ставка П2 в первом сете
                                recommendation = f"Ставка: П2 в первом сете"
                            else:
                                # Ничья в личных встречах
                                recommendation = "Ставка: Рекомендация недоступна (ничья в личных встречах)"
                        except Exception as e:
                            logger.error(
                                f"Ошибка при формировании рекомендации для игры {game['player1']} vs {game['player2']}: {e}")
                            recommendation = "Ставка: Ошибка при формировании рекомендации"
                        # --- Добавляем ссылку на игру (без преамбулы) ---
                        game_link_1xbet = game.get('game_link', None)
                        if game_link_1xbet:
                            link_message = f"\n<a href='{game_link_1xbet}'>Перейти к игре на 1xbet</a>"
                        else:
                            link_message = ""
                        message = (
                            f"🏆 <b>Анализ предстоящей игры</b>\n"
                            f"🏟️ <b>Лига:</b> {league_info}\n"
                            f"👤 <b>Игрок 1:</b> {game['player1']}\n"
                            f"👤 <b>Игрок 2:</b> {game['player2']}\n"
                            f"🕒 <b>Время начала:</b> {formatted_time}\n"
                            f"📈 <b>Личные встречи:</b> {game['last_encounters_score']}\n"
                            f"🔢 <b>Общая сумма очков:</b> {game['total_last_encounters_score']}\n"
                            f"💡 <b>{recommendation}</b>"  # Добавляем рекомендацию
                            f"{link_message}"  # Добавляем ссылку на игру БЕЗ преамбулы
                        )
                        # -----------------------------
                        await send_telegram_message(message)
                        mark_game_as_sent(game)
                        # --- Добавляем запись в Excel ---
                        append_game_to_excel(game) # <-- Вызов новой функции
                        # ----------------------------------
                        new_games_count += 1
                        logger.info(f"Отправлено сообщение для игры: {game['player1']} vs {game['player2']}")
                        logger.info(f"  Ссылка 1xbet: {game_link_1xbet}")
                        logger.info(f"  Рекомендация: {recommendation}")
                    else:
                        logger.debug(f"Игра уже была отправлена ранее: {game['player1']} vs {game['player2']}")
                else:
                    logger.debug(
                        f"Игра не прошла фильтр по сумме очков ({game['total_last_encounters_score']} < 10): {game['player1']} vs {game['player2']}")
            if new_games_count == 0:
                logger.info("Новых игр, подходящих под критерии (сумма >= 10), не найдено или все уже были отправлены.")
            else:
                logger.info(f"Успешно отправлено {new_games_count} новых сообщений.")
        else:
            logger.error(f"Ошибка получения страницы. Статус код: {response.status_code}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Ошибка сетевого запроса: {e}")
    except Exception as e:
        logger.error(f"Непредвиденная ошибка во время проверки игр: {e}", exc_info=True)

async def main():
    """Главная асинхронная функция бота."""
    logger.info("Бот запущен (локально).")
    # Загружаем состояние при старте
    load_sent_games()
    # --- Инициализируем Excel файл при старте ---
    initialize_excel_file() # <-- Вызов новой функции
    # --------------------------------------------
    while True:
        try:
            await check_and_send_games()
        except Exception as e:
            logger.error(f"Критическая ошибка в основном цикле: {e}", exc_info=True)
        finally:
            logger.info(f"Ожидаем {CHECK_INTERVAL_SECONDS} секунд до следующей проверки...")
            await asyncio.sleep(CHECK_INTERVAL_SECONDS)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Бот остановлен пользователем.")
    except Exception as e:
        logger.critical(f"Критическая ошибка при запуске бота: {e}", exc_info=True)
